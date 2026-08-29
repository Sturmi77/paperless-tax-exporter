"""
Tests fuer Kontoauszug-Matching (Issue #25 / #34).
"""
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ.setdefault("OUTPUT_DIR", tempfile.mkdtemp())
os.environ.setdefault("PAPERLESS_URL", "http://localhost:8000")
os.environ.setdefault("PAPERLESS_TOKEN", "test-token")
os.environ.setdefault("WINDOWS_UNC_PATH", "")

from bank_csv import parse_de_amount, parse_bank_csv, csv_preview, parse_iso_date  # noqa: E402
from matching_csv import (  # noqa: E402
    match_invoices_to_bank,
    amounts_match,
    STATUS_FOUND,
    STATUS_AMBIGUOUS,
    STATUS_NOT_FOUND,
    STATUS_NO_AMOUNT,
)
from excel_export import (  # noqa: E402
    create_excel,
    read_invoices_for_matching,
    update_excel_with_bank_matches,
    prepare_stb_export,
    COLOR_MATCH_OK,
    COLOR_MATCH_AMBIG,
    COLOR_MATCH_MISS,
    COLOR_MATCH_NO_AMT,
)
import openpyxl  # noqa: E402
from openpyxl.utils import range_boundaries  # noqa: E402

FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "at_giro_sample.csv")


class TestBankCsvParse:
    def test_parse_de_amount(self):
        assert parse_de_amount("-1.400,00") == -1400.0
        assert parse_de_amount("4.952,37") == 4952.37
        assert parse_de_amount("-16,05") == -16.05

    def test_parse_iso_date(self):
        assert parse_iso_date("2025-12-31").isoformat() == "2025-12-31"

    def test_preset_and_debits_only(self):
        rows = parse_bank_csv(FIXTURE, debits_only=True, only_relevant=False)
        assert all(r["amount"] is None or r["amount"] < 0 or not r["selected"] for r in rows)
        selected = [r for r in rows if r["selected"]]
        assert len(selected) == 3  # zwei -42,50 + eine -10,00; Gutschrift raus
        assert rows[0]["preset"] == "at_giro"

    def test_only_relevant(self):
        rows = parse_bank_csv(FIXTURE, debits_only=False, only_relevant=True)
        selected = [r for r in rows if r["selected"]]
        assert len(selected) == 1
        assert selected[0]["partner"] == "ACME GmbH"

    def test_preview(self):
        p = csv_preview(FIXTURE)
        assert p["preset"] == "at_giro"
        assert p["delimiter"] == ";"
        assert "Buchungsdatum" in p["fieldnames"]


class TestMatching:
    def test_amounts_match(self):
        assert amounts_match(42.5, -42.5)
        assert amounts_match(100.0, -100.01)
        assert not amounts_match(42.5, -10.0)

    def test_unique_match(self):
        from datetime import date
        invoices = [{
            "row": 5,
            "beleg_nr": 1,
            "re_dat": date(2025, 3, 10),
            "absender": "ACME GmbH",
            "beschreibung": "Rechnung Software",
            "betrag": 42.50,
        }]
        bank = parse_bank_csv(FIXTURE, debits_only=True)
        # nur erste ACME-Zeile (ohne Duplikat) fuer Eindeutigkeit
        bank = [b for b in bank if b["selected"] and "Duplikat" not in (b.get("text") or "")]
        result = match_invoices_to_bank(invoices, bank)
        assert result["stats"]["gefunden"] == 1
        assert result["matches"][0]["status"] == STATUS_FOUND
        assert result["matches"][0]["date"].isoformat() == "2025-03-15"

    def test_ambiguous_two_same_amount(self):
        from datetime import date
        invoices = [{
            "row": 5,
            "beleg_nr": 1,
            "re_dat": date(2025, 3, 10),
            "absender": "ACME",
            "beschreibung": "Rechnung",
            "betrag": 42.50,
        }]
        bank = parse_bank_csv(FIXTURE, debits_only=True)
        result = match_invoices_to_bank(invoices, bank, min_score=0.1)
        # zwei -42,50 ACME-Zeilen → volle Haerte: immer mehrdeutig
        assert result["stats"]["mehrdeutig"] == 1
        assert result["stats"]["gefunden"] == 0
        assert result["ambiguous"][0]["status"] == STATUS_AMBIGUOUS

    def test_no_blind_amount_only_match(self):
        """Ohne Textscore kein Treffer – auch bei eindeutigem Betrag."""
        from datetime import date
        invoices = [{
            "row": 5,
            "beleg_nr": 1,
            "re_dat": date(2025, 3, 10),
            "absender": "ZZZQQQ Fremdfirma 999",
            "beschreibung": "Komplett anderer Betreff xyzzy",
            "betrag": 42.50,
        }]
        bank = parse_bank_csv(FIXTURE, debits_only=True)
        # nur eine Betragszeile – frueher waere das ein Blind-Match gewesen
        bank = [b for b in bank if b["selected"] and "Duplikat" not in (b.get("text") or "")]
        result = match_invoices_to_bank(invoices, bank, min_score=0.25)
        assert result["stats"]["gefunden"] == 0
        assert result["stats"]["nicht_gefunden"] == 1
        assert result["unmatched"][0]["status"] == STATUS_NOT_FOUND

    def test_no_amount_status(self):
        from datetime import date
        invoices = [{
            "row": 5,
            "beleg_nr": 1,
            "re_dat": date(2025, 3, 10),
            "absender": "ACME",
            "beschreibung": "x",
            "betrag": None,
        }]
        bank = parse_bank_csv(FIXTURE, debits_only=True)
        result = match_invoices_to_bank(invoices, bank)
        assert result["stats"]["kein_betrag"] == 1
        assert result["no_amount"][0]["status"] == STATUS_NO_AMOUNT

    def test_one_to_one_bank_lock(self):
        from datetime import date
        invoices = [
            {
                "row": 5, "beleg_nr": 1, "re_dat": date(2025, 3, 10),
                "absender": "ACME GmbH", "beschreibung": "Rechnung Software", "betrag": 42.50,
            },
            {
                "row": 6, "beleg_nr": 2, "re_dat": date(2025, 3, 10),
                "absender": "ACME GmbH", "beschreibung": "Rechnung Software", "betrag": 42.50,
            },
        ]
        bank = parse_bank_csv(FIXTURE, debits_only=True)
        bank = [b for b in bank if b["selected"] and "Duplikat" not in (b.get("text") or "")]
        # nur eine 42,50-Zeile uebrig → erste Rechnung gefunden, zweite nicht/mehrdeutig
        result = match_invoices_to_bank(invoices, bank)
        assert result["stats"]["gefunden"] == 1
        assert result["stats"]["bank_used"] == 1
        assert result["stats"]["gefunden"] + result["stats"]["nicht_gefunden"] + result["stats"]["mehrdeutig"] == 2


class TestExcelBankUpdate:
    def test_read_only_empty_payment_date(self, tmp_path):
        from datetime import date
        docs = [
            {"id": 1, "title": "A", "created": "2025-03-10", "archive_serial_number": 1,
             "correspondent_name": "ACME GmbH"},
            {"id": 2, "title": "B", "created": "2025-03-11", "archive_serial_number": 2,
             "correspondent_name": "Other"},
        ]
        path = str(tmp_path / "Rechnungsaufstellung_2025.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 42.50
        ws.cell(row=6, column=8).value = 10.00
        ws.cell(row=6, column=3).value = date(2025, 3, 20)
        wb.save(path)
        inv = read_invoices_for_matching(path)
        assert len(inv) == 1
        assert inv[0]["beleg_nr"] == 1
        assert inv[0]["betrag"] == 42.50

    def test_betrag_prefers_column_i_full_amount(self, tmp_path):
        """Bank bucht Vollbetrag (I); H oft Anteil-Formel ohne Cache → None."""
        docs = [{
            "id": 1, "title": "Muell", "created": "2025-07-21",
            "archive_serial_number": 96, "correspondent_name": "Gemeindeverband",
        }]
        path = str(tmp_path / "privatanteil.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        # Wie in bearbeiteter STB-Datei: H = Anteil-Formel, I = Zahlungsbetrag
        ws.cell(row=5, column=8).value = (
            "=Tabelle1[[#This Row],[Rechnungs-\nsumme inkl. Privatanteil]]*0.11"
        )
        ws.cell(row=5, column=9).value = 185.66
        wb.save(path)
        inv = read_invoices_for_matching(path)
        assert len(inv) == 1
        assert inv[0]["betrag"] == 185.66

    def test_rows_without_beleg_nr_but_with_content(self, tmp_path):
        """Zeilen ohne A, aber mit Absender/Datum/Betrag, sind matchbar."""
        from datetime import date
        docs = [{
            "id": 1, "title": "Mit Beleg", "created": "2025-01-10",
            "archive_serial_number": 1, "correspondent_name": "ACME",
        }]
        path = str(tmp_path / "ohne_beleg.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 10.0
        # Extra-Zeile: kein Beleg-Nr., sonst voll
        ws.cell(row=6, column=1).value = None
        ws.cell(row=6, column=2).value = date(2025, 10, 24)
        ws.cell(row=6, column=5).value = "Stadtgemeinde Hollabrunn"
        ws.cell(row=6, column=6).value = "Wasser"
        ws.cell(row=6, column=9).value = 267.39
        # Padding-Zeile ohne Inhalt
        ws.cell(row=7, column=1).value = None
        wb.save(path)
        inv = read_invoices_for_matching(path)
        assert len(inv) == 2
        assert inv[0]["beleg_nr"] == 1
        assert inv[1]["beleg_nr"] is None
        assert inv[1]["absender"] == "Stadtgemeinde Hollabrunn"
        assert inv[1]["betrag"] == 267.39

    def test_collect_missing_beleg_issues(self, tmp_path):
        from datetime import date
        from excel_export import collect_missing_beleg_issues, plan_auto_beleg_numbers
        docs = [{
            "id": 1, "title": "Mit Beleg", "created": "2025-01-10",
            "archive_serial_number": 1, "correspondent_name": "ACME",
        }]
        path = str(tmp_path / "issues.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 10.0
        ws.cell(row=6, column=1).value = None
        ws.cell(row=6, column=2).value = date(2025, 10, 24)
        ws.cell(row=6, column=5).value = "Hollabrunn"
        ws.cell(row=6, column=9).value = 50.0
        wb.save(path)
        inv = read_invoices_for_matching(path)
        planned = plan_auto_beleg_numbers(path)
        issues = collect_missing_beleg_issues(inv, planned)
        assert len(issues) == 1
        assert issues[0]["row"] == 6
        assert issues[0]["proposed_beleg_nr"] == 2  # max existing = 1
        assert issues[0]["code"] == "beleg_auto_assign"

    def test_apply_auto_beleg_numbers_yellow(self, tmp_path):
        from datetime import date
        from excel_export import (
            apply_auto_beleg_numbers, plan_auto_beleg_numbers, COLOR_OCR_BG,
            BELEG_AUTO_COMMENT,
        )
        docs = [
            {"id": 1, "title": "A", "created": "2025-01-10",
             "archive_serial_number": 10, "correspondent_name": "ACME"},
            {"id": 2, "title": "B", "created": "2025-01-11",
             "archive_serial_number": 50, "correspondent_name": "Beta"},
        ]
        path = str(tmp_path / "auto_beleg.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 1.0
        ws.cell(row=6, column=8).value = 2.0
        # Dritte Zeile ohne Beleg
        ws.cell(row=7, column=1).value = None
        ws.cell(row=7, column=2).value = date(2025, 2, 1)
        ws.cell(row=7, column=5).value = "Neu GmbH"
        ws.cell(row=7, column=9).value = 99.0
        wb.save(path)

        planned = plan_auto_beleg_numbers(path)
        assert len(planned) == 1
        assert planned[0]["proposed_beleg_nr"] == 51  # max(10,50)+1

        applied = apply_auto_beleg_numbers(path, planned)
        assert len(applied) == 1
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2["Rechnungsaufstellung"]
        cell = ws2.cell(row=7, column=1)
        assert cell.value == 51
        assert cell.fill.fgColor.rgb[-6:].upper() == COLOR_OCR_BG
        assert cell.comment is not None
        assert BELEG_AUTO_COMMENT in (cell.comment.text or "")

    def test_auto_beleg_does_not_overwrite_existing_or_formula(self, tmp_path):
        """Schreibschutz: vorhandene A-Werte und Formeln bleiben."""
        from datetime import date
        from excel_export import apply_auto_beleg_numbers, plan_auto_beleg_numbers
        docs = [{
            "id": 1, "title": "A", "created": "2025-01-10",
            "archive_serial_number": 5, "correspondent_name": "ACME",
        }]
        path = str(tmp_path / "preserve_a.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 1.0
        # Zeile mit Formel in A
        ws.cell(row=6, column=1).value = "=5+1"
        ws.cell(row=6, column=2).value = date(2025, 2, 1)
        ws.cell(row=6, column=5).value = "Formel GmbH"
        ws.cell(row=6, column=9).value = 10.0
        # Zeile mit Platzhalter (gilt als vorhandener Wert)
        ws.cell(row=7, column=1).value = "-"
        ws.cell(row=7, column=2).value = date(2025, 2, 2)
        ws.cell(row=7, column=5).value = "Platzhalter AG"
        ws.cell(row=7, column=9).value = 11.0
        # Wirklich leere A → darf vergeben werden
        ws.cell(row=8, column=1).value = None
        ws.cell(row=8, column=2).value = date(2025, 2, 3)
        ws.cell(row=8, column=5).value = "Neu GmbH"
        ws.cell(row=8, column=9).value = 12.0
        wb.save(path)

        planned = plan_auto_beleg_numbers(path)
        assert len(planned) == 1
        assert planned[0]["row"] == 8
        assert planned[0]["proposed_beleg_nr"] == 6  # max=5

        applied = apply_auto_beleg_numbers(path, planned)
        assert len(applied) == 1
        assert applied[0]["beleg_nr"] == 6

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2["Rechnungsaufstellung"]
        assert ws2.cell(row=5, column=1).value == 5
        assert ws2.cell(row=6, column=1).value == "=5+1"
        assert ws2.cell(row=7, column=1).value == "-"
        assert ws2.cell(row=8, column=1).value == 6

    def test_update_colors_c_and_preserves_formulas(self, tmp_path):
        from datetime import date
        docs = [{
            "id": 1, "title": "Software", "created": "2025-03-10",
            "archive_serial_number": 7, "correspondent_name": "ACME GmbH",
        }]
        path = str(tmp_path / "Rechnungsaufstellung_2025.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 42.50
        ws.cell(row=5, column=9).value = "=H5*0.5"
        sum_before = ws["H1"].value
        table_ref_before = list(ws.tables.values())[0].ref
        min_c, min_r, max_c, max_r = range_boundaries(table_ref_before)
        wb.save(path)

        bank = parse_bank_csv(FIXTURE, debits_only=True)
        bank = [b for b in bank if b["selected"] and "Duplikat" not in (b.get("text") or "")]
        invoices = read_invoices_for_matching(path)
        result = match_invoices_to_bank(invoices, bank)
        assert result["stats"]["gefunden"] == 1
        n = update_excel_with_bank_matches(path, result)
        assert n >= 1

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2["Rechnungsaufstellung"]
        assert ws2["H1"].value == sum_before
        assert ws2.cell(row=5, column=9).value == "=H5*0.5"
        assert ws2.cell(row=5, column=3).value is not None
        assert ws2.cell(row=5, column=3).fill.fgColor.rgb[-6:].upper() == COLOR_MATCH_OK
        headers = [ws2.cell(row=4, column=c).value for c in range(1, ws2.max_column + 1)]
        assert any(h and "Match-Status" in str(h) for h in headers)
        ref_after = list(ws2.tables.values())[0].ref
        _c1, r1, _c2, r2 = range_boundaries(ref_after)
        assert r1 == min_r and r2 == max_r

    def test_ambiguous_and_miss_color_c(self, tmp_path):
        from datetime import date
        docs = [
            {"id": 1, "title": "A", "created": "2025-03-10", "archive_serial_number": 1,
             "correspondent_name": "ACME"},
            {"id": 2, "title": "B", "created": "2025-03-10", "archive_serial_number": 2,
             "correspondent_name": "Nobody"},
        ]
        path = str(tmp_path / "colors.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 42.50
        ws.cell(row=6, column=8).value = 999.99
        wb.save(path)

        bank = parse_bank_csv(FIXTURE, debits_only=True)
        invoices = read_invoices_for_matching(path)
        result = match_invoices_to_bank(invoices, bank, min_score=0.1)
        update_excel_with_bank_matches(path, result)

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2["Rechnungsaufstellung"]
        # Zeile 5 mehrdeutig (2x 42,50) → C leer + gelb
        assert ws2.cell(row=5, column=3).value in (None, "")
        assert ws2.cell(row=5, column=3).fill.fgColor.rgb[-6:].upper() == COLOR_MATCH_AMBIG
        # Zeile 6 nicht gefunden → C leer + rot
        assert ws2.cell(row=6, column=3).value in (None, "")
        assert ws2.cell(row=6, column=3).fill.fgColor.rgb[-6:].upper() == COLOR_MATCH_MISS

    def test_no_amount_gray_c(self, tmp_path):
        docs = [{
            "id": 1, "title": "NoAmt", "created": "2025-03-10",
            "archive_serial_number": 1, "correspondent_name": "X",
        }]
        path = str(tmp_path / "noamt.xlsx")
        create_excel(docs, {}, path, "2025")
        # H leer lassen
        bank = parse_bank_csv(FIXTURE, debits_only=True)
        invoices = read_invoices_for_matching(path)
        result = match_invoices_to_bank(invoices, bank)
        assert result["stats"]["kein_betrag"] == 1
        update_excel_with_bank_matches(path, result)
        wb = openpyxl.load_workbook(path)
        cell = wb["Rechnungsaufstellung"].cell(row=5, column=3)
        assert cell.value in (None, "")
        assert cell.fill.fgColor.rgb[-6:].upper() == COLOR_MATCH_NO_AMT

    def test_formula_in_c_not_overwritten(self, tmp_path):
        docs = [{
            "id": 1, "title": "Software", "created": "2025-03-10",
            "archive_serial_number": 7, "correspondent_name": "ACME GmbH",
        }]
        path = str(tmp_path / "with_formula_c.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 42.50
        wb.save(path)

        bank = parse_bank_csv(FIXTURE, debits_only=True)
        bank = [b for b in bank if b["selected"] and "Duplikat" not in (b.get("text") or "")]
        invoices = read_invoices_for_matching(path)
        result = match_invoices_to_bank(invoices, bank)

        wb2 = openpyxl.load_workbook(path)
        wb2["Rechnungsaufstellung"].cell(row=5, column=3).value = "=B5"
        wb2.save(path)

        update_excel_with_bank_matches(path, result)
        wb3 = openpyxl.load_workbook(path)
        assert wb3["Rechnungsaufstellung"].cell(row=5, column=3).value == "=B5"

    def test_stb_export_clears_styles_keeps_date(self, tmp_path):
        from datetime import date
        docs = [{
            "id": 1, "title": "Software", "created": "2025-03-10",
            "archive_serial_number": 7, "correspondent_name": "ACME GmbH",
        }]
        path = str(tmp_path / "stb.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        wb["Rechnungsaufstellung"].cell(row=5, column=8).value = 42.50
        wb.save(path)

        bank = parse_bank_csv(FIXTURE, debits_only=True)
        bank = [b for b in bank if b["selected"] and "Duplikat" not in (b.get("text") or "")]
        invoices = read_invoices_for_matching(path)
        result = match_invoices_to_bank(invoices, bank)
        update_excel_with_bank_matches(path, result)

        pay_before = openpyxl.load_workbook(path)["Rechnungsaufstellung"].cell(row=5, column=3).value
        assert pay_before is not None

        stats = prepare_stb_export(path)
        assert stats["status_column_hidden"] is True

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2["Rechnungsaufstellung"]
        cell_c = ws2.cell(row=5, column=3)
        assert cell_c.value == pay_before
        assert cell_c.fill.fill_type is None
        assert cell_c.comment is None
        headers = {
            (ws2.cell(row=4, column=c).value or "").strip(): c
            for c in range(1, ws2.max_column + 1)
        }
        col_s = headers.get("Match-Status")
        assert col_s
        assert ws2.cell(row=5, column=col_s).value in (None, "")
        assert ws2.column_dimensions[openpyxl.utils.get_column_letter(col_s)].hidden is True

    def test_only_empty_cells_get_values(self, tmp_path):
        """Befuellte C / Buchungstext / manueller Status bleiben unangetastet."""
        from datetime import date
        from excel_export import _ensure_bank_match_columns
        docs = [{
            "id": 1, "title": "Software", "created": "2025-03-10",
            "archive_serial_number": 7, "correspondent_name": "ACME GmbH",
        }]
        path = str(tmp_path / "preserve.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 42.50
        # Match-Spalten anlegen und manuell befuellen, C leer lassen fuer Read
        col_text, col_status = _ensure_bank_match_columns(ws)
        ws.cell(row=5, column=col_text).value = "MANUELLER TEXT"
        ws.cell(row=5, column=col_status).value = "bitte pruefen"
        wb.save(path)

        bank = parse_bank_csv(FIXTURE, debits_only=True)
        bank = [b for b in bank if b["selected"] and "Duplikat" not in (b.get("text") or "")]
        invoices = read_invoices_for_matching(path)
        result = match_invoices_to_bank(invoices, bank)
        assert result["stats"]["gefunden"] == 1
        update_excel_with_bank_matches(path, result)

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2["Rechnungsaufstellung"]
        assert ws2.cell(row=5, column=3).value is not None  # C war leer → Datum ok
        assert ws2.cell(row=5, column=col_text).value == "MANUELLER TEXT"
        assert ws2.cell(row=5, column=col_status).value == "bitte pruefen"

    def test_filled_c_not_overwritten_even_if_in_result(self, tmp_path):
        from datetime import date
        docs = [{
            "id": 1, "title": "Software", "created": "2025-03-10",
            "archive_serial_number": 7, "correspondent_name": "ACME GmbH",
        }]
        path = str(tmp_path / "filled_c.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 42.50
        existing = date(2025, 1, 1)
        ws.cell(row=5, column=3).value = existing
        wb.save(path)

        # Ergebnis kuenstlich (Read wuerde Zeile ueberspringen)
        fake = {
            "matches": [{
                "invoice_row": 5,
                "date": date(2025, 3, 15),
                "text": "soll nicht schreiben",
                "status": STATUS_FOUND,
            }],
            "ambiguous": [],
            "unmatched": [],
            "no_amount": [],
        }
        update_excel_with_bank_matches(path, fake)
        wb2 = openpyxl.load_workbook(path)
        got = wb2["Rechnungsaufstellung"].cell(row=5, column=3).value
        if hasattr(got, "date"):
            got = got.date()
        assert got == existing
        # kein Match-Datum 2025-03-15 geschrieben
        assert got != date(2025, 3, 15)
    def test_table_columns_synced_no_second_table_expand(self, tmp_path):
        """
        Ref-Erweiterung muss tableColumns syncen; zweite Tabelle nicht anfassen
        (sonst Excel-Reparatur: table2.xml entfernt).
        """
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import range_boundaries

        docs = [{
            "id": 1, "title": "Software", "created": "2025-03-10",
            "archive_serial_number": 7, "correspondent_name": "ACME GmbH",
        }]
        path = str(tmp_path / "two_tables.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        ws = wb["Rechnungsaufstellung"]
        ws.cell(row=5, column=8).value = 42.50
        # Zweite Tabelle unterhalb, ohne Header-Zeile 4 (kein Einfluss auf next_col)
        ws.cell(row=10, column=20).value = "Notiz"
        ws.cell(row=11, column=20).value = "x"
        t2 = Table(displayName="Hilfstabelle", ref="T10:T11")
        t2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(t2)
        t1_before = [t for t in ws.tables.values() if t.displayName == "Tabelle1"][0]
        t1_ref_before = t1_before.ref
        t2_ref_before = "T10:T11"
        wb.save(path)

        bank = parse_bank_csv(FIXTURE, debits_only=True)
        bank = [b for b in bank if b["selected"] and "Duplikat" not in (b.get("text") or "")]
        invoices = read_invoices_for_matching(path)
        result = match_invoices_to_bank(invoices, bank)
        update_excel_with_bank_matches(path, result)

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2["Rechnungsaufstellung"]
        assert "Tabelle1" in [t.displayName for t in ws2.tables.values()]
        assert "Hilfstabelle" in [t.displayName for t in ws2.tables.values()]
        t1 = [t for t in ws2.tables.values() if t.displayName == "Tabelle1"][0]
        t2b = [t for t in ws2.tables.values() if t.displayName == "Hilfstabelle"][0]
        assert t2b.ref == t2_ref_before
        _c1, r1, c2, r2 = range_boundaries(t1.ref)
        _a, _b, c2_before, _d = range_boundaries(t1_ref_before)
        assert c2 > c2_before
        assert r1 == 4 and r2 >= 5
        assert len(t1.tableColumns) == (c2 - _c1 + 1)
        names = [c.name for c in t1.tableColumns]
        assert len(names) == len(set(n.lower() for n in names))
        assert any("Match-Status" in n or "match" in n.lower() for n in names)
