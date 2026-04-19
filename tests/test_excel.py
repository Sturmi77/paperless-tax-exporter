"""
Tests für excel_export.py – UNC-Pfad-Logik und Subfolder-Schnittstelle.
Erweitert in Schritt 4 (Issue #8) um CELL-Formel-Tests.
"""
import pytest
import sys
import os
import tempfile

# Projektpfad
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel_export import _build_unc_path, _build_cell_formula, create_excel, append_to_excel


# ---------------------------------------------------------------------------
# Tests: _build_unc_path()
# ---------------------------------------------------------------------------

class TestBuildUncPath:
    UNC_BASE   = r"\\SynologyDS923\downloads\steuerberater"
    YEAR       = "2024"
    FILENAME   = "0012_Telekom.pdf"

    def test_standard_path_no_subfolder(self):
        result = _build_unc_path(self.UNC_BASE, self.YEAR, self.FILENAME)
        assert result == r"\\SynologyDS923\downloads\steuerberater\2024\Belege\0012_Telekom.pdf"

    def test_path_with_subfolder(self):
        result = _build_unc_path(self.UNC_BASE, self.YEAR, self.FILENAME, subfolder="Archiv")
        assert result == r"\\SynologyDS923\downloads\steuerberater\2024\Archiv\Belege\0012_Telekom.pdf"

    def test_empty_subfolder_equals_no_subfolder(self):
        result_empty   = _build_unc_path(self.UNC_BASE, self.YEAR, self.FILENAME, subfolder="")
        result_default = _build_unc_path(self.UNC_BASE, self.YEAR, self.FILENAME)
        assert result_empty == result_default

    def test_trailing_backslash_in_base_stripped(self):
        base_with_slash = self.UNC_BASE + "\\"
        result = _build_unc_path(base_with_slash, self.YEAR, self.FILENAME)
        # Kein doppelter Backslash
        assert "\\\\" not in result.lstrip("\\\\")

    def test_no_unc_base_returns_filename(self):
        result = _build_unc_path(None, self.YEAR, self.FILENAME)
        assert result == self.FILENAME

    def test_no_filename_returns_empty(self):
        result = _build_unc_path(self.UNC_BASE, self.YEAR, "")
        assert result == ""

    def test_both_none_returns_empty(self):
        result = _build_unc_path(None, self.YEAR, None)
        assert result == ""

    def test_subfolder_with_dash_and_digits(self):
        result = _build_unc_path(self.UNC_BASE, self.YEAR, self.FILENAME, subfolder="2024-Q4")
        assert "2024-Q4" in result
        assert r"\Belege\\" not in result  # kein doppelter Separator


# ---------------------------------------------------------------------------
# Tests: create_excel() – Grundfunktion + subfolder-Parameter
# ---------------------------------------------------------------------------

class TestCreateExcel:
    SAMPLE_DOCS = [
        {
            "id": 1,
            "archive_serial_number": "0001",
            "created": "2024-03-15",
            "correspondent_name": "Telekom",
            "title": "Telefonrechnung März",
            "document_type": 2,
            "document_type_name": "Rechnung",
        },
        {
            "id": 2,
            "archive_serial_number": "0002",
            "created": "2024-06-01",
            "correspondent_name": None,
            "title": "Internetrechnung",
            "document_type": 2,
            "document_type_name": "Rechnung",
        },
    ]
    PDF_MAP = {1: "0001_Telekom.pdf", 2: "0002_Internet.pdf"}
    UNC_BASE = r"\\SynologyDS923\downloads\steuerberater"

    def test_creates_file(self, tmp_path):
        out = str(tmp_path / "test.xlsx")
        result = create_excel(self.SAMPLE_DOCS, self.PDF_MAP, out, "2024")
        assert os.path.exists(result)
        assert os.path.getsize(result) > 0

    def test_creates_file_with_unc(self, tmp_path):
        out = str(tmp_path / "test_unc.xlsx")
        result = create_excel(
            self.SAMPLE_DOCS, self.PDF_MAP, out, "2024",
            unc_base=self.UNC_BASE
        )
        assert os.path.exists(result)

    def test_creates_file_with_subfolder(self, tmp_path):
        out = str(tmp_path / "test_subfolder.xlsx")
        result = create_excel(
            self.SAMPLE_DOCS, self.PDF_MAP, out, "2024",
            unc_base=self.UNC_BASE, subfolder="Archiv"
        )
        assert os.path.exists(result)

    def test_subfolder_appears_in_hyperlink(self, tmp_path):
        """Sicherstellen dass der Subfolder im Hyperlink-Pfad vorkommt."""
        import openpyxl
        out = str(tmp_path / "test_hyperlink.xlsx")
        create_excel(
            self.SAMPLE_DOCS, self.PDF_MAP, out, "2024",
            unc_base=self.UNC_BASE, subfolder="Archiv"
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["Rechnungsaufstellung"]
        # Spalte J ab Zeile 5 – Hyperlink-Formel prüfen
        cell_j = ws.cell(row=5, column=10)
        assert cell_j.value is not None
        assert "Archiv" in str(cell_j.value)
        wb.close()

    def test_no_subfolder_no_subfolder_in_hyperlink(self, tmp_path):
        """Ohne Subfolder: UNC-Modus muss Jahr + Belege im Pfad enthalten."""
        import openpyxl
        out = str(tmp_path / "test_no_sub.xlsx")
        create_excel(
            self.SAMPLE_DOCS, self.PDF_MAP, out, "2024",
            unc_base=self.UNC_BASE,
            hyperlink_mode="unc",
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["Rechnungsaufstellung"]
        cell_j = ws.cell(row=5, column=10)
        val = str(cell_j.value or "")
        # UNC-Pfad muss Jahr und Belege-Ordner enthalten, aber keinen doppelten Separator
        assert r"\2024\Belege\\" not in val  # kein doppelter Separator
        assert "2024" in val
        assert "Belege" in val
        wb.close()


# ---------------------------------------------------------------------------
# Tests: append_to_excel() – subfolder-Parameter
# ---------------------------------------------------------------------------

class TestAppendToExcel:
    INITIAL_DOCS = [
        {"id": 1, "archive_serial_number": "0001", "created": "2024-01-10",
         "correspondent_name": "Firma A", "title": "Rechnung 1",
         "document_type": 1, "document_type_name": "Rechnung"},
    ]
    NEW_DOCS = [
        {"id": 2, "archive_serial_number": "0002", "created": "2024-07-20",
         "correspondent_name": "Firma B", "title": "Rechnung 2",
         "document_type": 1, "document_type_name": "Rechnung"},
    ]
    PDF_MAP = {1: "0001.pdf", 2: "0002.pdf"}
    UNC_BASE = r"\\SynologyDS923\downloads\steuerberater"

    def test_append_adds_row(self, tmp_path):
        out = str(tmp_path / "append_test.xlsx")
        create_excel(self.INITIAL_DOCS, self.PDF_MAP, out, "2024", unc_base=self.UNC_BASE)
        added = append_to_excel(self.NEW_DOCS, self.PDF_MAP, out, "2024", unc_base=self.UNC_BASE)
        assert added == 1

    def test_append_with_subfolder(self, tmp_path):
        out = str(tmp_path / "append_subfolder.xlsx")
        create_excel(self.INITIAL_DOCS, self.PDF_MAP, out, "2024",
                     unc_base=self.UNC_BASE, subfolder="Q1")
        added = append_to_excel(self.NEW_DOCS, self.PDF_MAP, out, "2024",
                                 unc_base=self.UNC_BASE, subfolder="Q1")
        assert added == 1

    def test_append_empty_list_returns_zero(self, tmp_path):
        out = str(tmp_path / "append_empty.xlsx")
        create_excel(self.INITIAL_DOCS, self.PDF_MAP, out, "2024")
        added = append_to_excel([], self.PDF_MAP, out, "2024")
        assert added == 0


# ---------------------------------------------------------------------------
# Tests: _build_cell_formula() (Issue #8)
# ---------------------------------------------------------------------------

class TestBuildCellFormula:
    def test_formula_starts_with_hyperlink(self):
        formula = _build_cell_formula(r"Belege\test.pdf")
        assert formula.startswith("=HYPERLINK(")

    def test_formula_contains_cell_filename(self):
        formula = _build_cell_formula(r"Belege\test.pdf")
        assert 'CELL("filename")' in formula

    def test_formula_display_name_is_filename(self):
        formula = _build_cell_formula(r"Belege\test.pdf")
        assert '"test.pdf"' in formula

    def test_formula_with_subfolder(self):
        formula = _build_cell_formula(r"Archiv\Belege\test.pdf")
        assert 'CELL("filename")' in formula
        assert '"test.pdf"' in formula

    def test_formula_is_dynamic_no_static_path(self):
        """Wichtigste Eigenschaft: kein statischer absoluter Pfad eingebettet."""
        formula = _build_cell_formula(r"Belege\test.pdf")
        assert "C:\\" not in formula
        assert "\\\\SynologyDS923" not in formula


# ---------------------------------------------------------------------------
# Tests: create_excel() – CELL-Formel vs. UNC (Issue #8)
# ---------------------------------------------------------------------------

class TestCellFormulaHyperlinks:
    SAMPLE_DOCS = [
        {"id": 1, "archive_serial_number": "0001", "created": "2024-01-10",
         "correspondent_name": "Firma A", "title": "Rechnung 1",
         "document_type": 1, "document_type_name": "Rechnung"},
    ]
    PDF_MAP = {1: "0001_Test.pdf"}
    UNC_BASE = r"\\SynologyDS923\downloads\steuerberater"

    def test_hyperlink_mode_cell_generates_formula(self, tmp_path):
        """hyperlink_mode='cell' muss CELL("filename")-Formel in Spalte J erzeugen."""
        import openpyxl
        out = str(tmp_path / "cell_mode.xlsx")
        create_excel(self.SAMPLE_DOCS, self.PDF_MAP, out, "2024",
                     unc_base=self.UNC_BASE, hyperlink_mode="cell")
        wb = openpyxl.load_workbook(out)
        ws = wb["Rechnungsaufstellung"]
        cell_j = ws.cell(row=5, column=10)
        val = str(cell_j.value or "")
        assert 'CELL("filename")' in val
        assert "=HYPERLINK(" in val
        wb.close()

    def test_hyperlink_mode_unc_generates_absolute_path(self, tmp_path):
        """hyperlink_mode='unc' muss absoluten UNC-Pfad erzeugen (kein CELL)."""
        import openpyxl
        out = str(tmp_path / "unc_mode.xlsx")
        create_excel(self.SAMPLE_DOCS, self.PDF_MAP, out, "2024",
                     unc_base=self.UNC_BASE, hyperlink_mode="unc")
        wb = openpyxl.load_workbook(out)
        ws = wb["Rechnungsaufstellung"]
        cell_j = ws.cell(row=5, column=10)
        val = str(cell_j.value or "")
        assert "SynologyDS923" in val
        assert 'CELL("filename")' not in val
        wb.close()

    def test_include_text_path_adds_column_k(self, tmp_path):
        """include_text_path=True muss Spalte K mit UNC-Pfad erzeugen."""
        import openpyxl
        out = str(tmp_path / "text_path.xlsx")
        create_excel(self.SAMPLE_DOCS, self.PDF_MAP, out, "2024",
                     unc_base=self.UNC_BASE, include_text_path=True)
        wb = openpyxl.load_workbook(out)
        ws = wb["Rechnungsaufstellung"]
        # Header Zeile 4, Spalte K
        header_k = ws.cell(row=4, column=11).value
        assert header_k is not None and "Pfad" in str(header_k)
        # Daten Zeile 5, Spalte K
        data_k = ws.cell(row=5, column=11).value
        assert data_k is not None
        assert "SynologyDS923" in str(data_k)
        wb.close()

    def test_no_text_path_column_k_empty(self, tmp_path):
        """include_text_path=False (default) darf keine Spalte K erzeugen."""
        import openpyxl
        out = str(tmp_path / "no_text_path.xlsx")
        create_excel(self.SAMPLE_DOCS, self.PDF_MAP, out, "2024",
                     unc_base=self.UNC_BASE, include_text_path=False)
        wb = openpyxl.load_workbook(out)
        ws = wb["Rechnungsaufstellung"]
        data_k = ws.cell(row=5, column=11).value
        assert data_k is None
        wb.close()

    def test_cell_formula_no_static_path_embedded(self, tmp_path):
        """Wichtigste Eigenschaft: CELL-Formel enthält keinen statischen Server-Pfad."""
        import openpyxl
        out = str(tmp_path / "no_static.xlsx")
        create_excel(self.SAMPLE_DOCS, self.PDF_MAP, out, "2024",
                     unc_base=self.UNC_BASE, hyperlink_mode="cell")
        wb = openpyxl.load_workbook(out)
        ws = wb["Rechnungsaufstellung"]
        cell_j = ws.cell(row=5, column=10)
        val = str(cell_j.value or "")
        # Kein statischer NAS-Pfad in der Formel
        assert "SynologyDS923" not in val
        wb.close()
