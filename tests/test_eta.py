"""
Tests fuer ETA-Hilfsfunktion und Progress-Callbacks (Issue #19).
"""
import sys
import os
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ.setdefault("OUTPUT_DIR", "/tmp/pte-test-output")
os.environ.setdefault("PAPERLESS_URL", "http://localhost:8000")
os.environ.setdefault("PAPERLESS_TOKEN", "test-token")
os.environ.setdefault("WINDOWS_UNC_PATH", "")

from app import _calc_eta  # noqa: E402
from excel_export import create_excel  # noqa: E402
from pdf_export import download_pdfs  # noqa: E402


class TestCalcEta:
    def test_none_when_no_start(self):
        assert _calc_eta(None, None, 1, 10) == (None, None)

    def test_none_when_done_zero(self):
        t0 = time.monotonic()
        assert _calc_eta(t0, t0 + 1, 0, 10) == (None, None)

    def test_none_when_total_zero(self):
        t0 = time.monotonic()
        assert _calc_eta(t0, t0 + 1, 5, 0) == (None, None)

    def test_avg_and_eta(self):
        t0 = 1000.0
        last = 1020.0  # 20 s fuer 4 Docs → avg 5.0
        avg, eta = _calc_eta(t0, last, 4, 10)
        assert avg == 5.0
        assert eta == 30  # 5 * 6 remaining

    def test_eta_zero_when_done(self):
        t0 = 1000.0
        last = 1010.0
        avg, eta = _calc_eta(t0, last, 5, 5)
        assert avg == 2.0
        assert eta == 0


class TestInvoiceSheetFallback:
    def test_renamed_sheet_still_readable(self, tmp_path):
        from excel_export import create_excel, read_invoices_for_matching, _get_invoice_sheet
        import openpyxl
        docs = [{
            "id": 1, "title": "A", "created": "2025-03-10",
            "archive_serial_number": 1, "correspondent_name": "ACME",
        }]
        path = str(tmp_path / "edited.xlsx")
        create_excel(docs, {}, path, "2025")
        wb = openpyxl.load_workbook(path)
        wb["Rechnungsaufstellung"].cell(row=5, column=8).value = 12.5
        # Sheet umbenennen wie bei manueller Bearbeitung
        wb["Rechnungsaufstellung"].title = "Tabelle1"
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        assert "Rechnungsaufstellung" not in wb2.sheetnames
        ws = _get_invoice_sheet(wb2)
        assert ws.title == "Tabelle1"

        inv = read_invoices_for_matching(path)
        assert len(inv) == 1
        assert inv[0]["betrag"] == 12.5


class TestCreateExcelProgress:
    def test_progress_fn_called_per_doc(self, tmp_path):
        docs = [
            {"id": 1, "title": "A", "created": "2024-01-01", "archive_serial_number": 1},
            {"id": 2, "title": "B", "created": "2024-01-02", "archive_serial_number": 2},
            {"id": 3, "title": "C", "created": "2024-01-03", "archive_serial_number": 3},
        ]
        calls = []

        def progress(idx, total, title):
            calls.append((idx, total, title))

        out = tmp_path / "test.xlsx"
        create_excel(docs, {}, str(out), "2024", progress_fn=progress)
        assert len(calls) == 3
        assert calls[0] == (1, 3, "A")
        assert calls[2][0] == 3 and calls[2][1] == 3


class TestDownloadPdfsProgress:
    def test_progress_fn_called_even_when_file_exists(self, tmp_path):
        # Vorhandene Datei → Download uebersprungen, Progress trotzdem gemeldet
        doc = {
            "id": 42,
            "title": "Beleg_Test",
            "archive_serial_number": 7,
        }
        # Dateiname wie _make_pdf_filename: 0007_Beleg_Test.pdf
        existing = tmp_path / "0007_Beleg_Test.pdf"
        existing.write_bytes(b"%PDF-1.4")

        calls = []

        def progress(idx, total, title):
            calls.append((idx, total, title))

        result = download_pdfs(
            [doc], str(tmp_path), "http://localhost:8000", "tok",
            progress_fn=progress,
        )
        assert len(calls) == 1
        assert calls[0] == (1, 1, "Beleg_Test")
        assert 42 in result
