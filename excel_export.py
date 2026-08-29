"""
Excel-Export im Format der Steuerberater-Vorlage.
Struktur: Tabelle1 mit 10 Spalten, Header in Zeile 4, SUMME in Zeile 1.
Spalten:
  A: Beleg-Nr.
  B: Re-Dat
  C: Zahlungsdatum  (leer – manuell zu befüllen)
  D: Bar / Konto / KK (leer – manuell zu befüllen)
  E: Absender (correspondent aus Paperless oder OCR-Vorschlag gelb)
  F: Beschreibung
  G: Kennzahl (document_type)
  H: Rechnungssumme (leer Stufe 1 / OCR-Vorschlag Stufe 2 gelb)
  I: Rechnungssumme inkl. Privatanteil (leer – manuell)
  J: Dateiname (Hyperlink zur PDF)
"""

import os
import re
from datetime import datetime, date
from typing import Optional
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# Farben – komm|event CI
COLOR_HEADER_BG   = "2997AB"   # Teal (komm|event Primärfarbe)
COLOR_HEADER_FONT = "FFFFFF"   # Weiß
COLOR_SUM_BG      = "D6EEF2"   # Teal-hell (abgeleitet)
COLOR_SUM_FONT    = "1A6478"   # Teal-dunkel für SUMME-Beschriftung
COLOR_OCR_BG      = "FFFFC7"   # Gelb für OCR-Vorschlagswerte (funktional, bleibt)
COLOR_EMPTY_BG    = "F2F2F2"   # Hellgrau für manuell zu füllende Felder
COLOR_HYPERLINK   = "1E7D8F"   # Teal-dunkel für Hyperlinks

THIN   = Side(style="thin", color="C5D2D4")  # CI-Border
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Beleg-Nr.",                           5.8,   "center"),
    ("Re-Dat",                              12.0,  "center"),
    ("Zahlungs-\ndatum",                    12.0,  "center"),
    ("Bar / Konto / KK",                    13.0,  "center"),
    ("Absender",                            28.0,  "left"),
    ("Beschreibung",                        38.0,  "left"),
    ("Kennzahl",                            22.0,  "left"),
    ("Rechnungssumme",                      20.0,  "right"),
    ("Rechnungs-\nsumme inkl. Privatanteil", 18.0, "right"),
    ("Dateiname / Beleg",                   38.0,  "left"),
]

# Optionale Spalte K (INCLUDE_TEXT_PATH=true)
COLUMN_TEXT_PATH = ("Pfad (kopierbar)", 52.0, "left")

# Issue #25: Bank-Match-Spalten (werden bei Bedarf angehängt)
COLUMN_BOOKING_TEXT = ("Buchungstext", 42.0, "left")
COLUMN_MATCH_STATUS = ("Match-Status", 16.0, "center")

COLOR_MATCH_OK      = "C6EFCE"  # Grün – gefunden
COLOR_MATCH_AMBIG   = "FFFFC7"  # Gelb – mehrdeutig
COLOR_MATCH_MISS    = "FFC7CE"  # Rot – nicht gefunden (Issue #34)
COLOR_MATCH_NO_AMT  = "D9D9D9"  # Grau – kein Betrag

DATE_FORMAT   = "DD.MM.YYYY"
NUMBER_FORMAT = '#,##0.00 "€"'

SHEET_INVOICE = "Rechnungsaufstellung"


def _get_invoice_sheet(wb):
    """
    Liefert das Rechnungs-Worksheet.
    1) Exakter Name „Rechnungsaufstellung“
    2) Sheet mit Header-Zeile 4 (Beleg-Nr. / Re-Dat)
    3) Erstes Worksheet
    """
    if SHEET_INVOICE in wb.sheetnames:
        return wb[SHEET_INVOICE]

    def _norm(v):
        if v is None:
            return ""
        return str(v).replace("\n", " ").strip().lower()

    for name in wb.sheetnames:
        ws = wb[name]
        headers = [_norm(ws.cell(row=4, column=c).value) for c in range(1, 6)]
        joined = " ".join(headers)
        if "beleg" in joined or "re-dat" in joined or "re dat" in joined:
            return ws

    if wb.sheetnames:
        return wb[wb.sheetnames[0]]

    raise ValueError(
        "Excel enthält kein Worksheet. "
        f"Vorhanden: {wb.sheetnames!r}"
    )


def _header_cell(ws, row, col, value, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
    cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = BORDER
    return cell


def _data_cell(ws, row, col, value=None, align="left", number_fmt=None,
               bold=False, bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(size=10, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = BORDER
    if number_fmt:
        cell.number_format = number_fmt
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _make_comment(text):
    try:
        from openpyxl.comments import Comment
        return Comment(text, "Paperless Exporter")
    except Exception:
        return None


def _build_unc_path(unc_base, year_label, filename, subfolder: str = ""):
    """
    Baut Windows-UNC-Pfad für Excel-Hyperlink.
    unc_base:   z.B. \\\\SynologyDS923\\downloads\\steuerberater
    year_label: z.B. 2024
    filename:   z.B. 0012_Telekom.pdf
    subfolder:  optionaler Unterordner (Allowlist-validiert, default="")
    Ergebnis (ohne subfolder): \\\\SynologyDS923\\downloads\\steuerberater\\2024\\Belege\\0012_Telekom.pdf
    Ergebnis (mit subfolder):  \\\\SynologyDS923\\downloads\\steuerberater\\2024\\Archiv\\Belege\\0012_Telekom.pdf
    """
    if not unc_base or not filename:
        return filename or ""
    # Backslashes normalisieren
    base = unc_base.rstrip("\\")
    if subfolder:
        return f"{base}\\{year_label}\\{subfolder}\\Belege\\{filename}"
    return f"{base}\\{year_label}\\Belege\\{filename}"


def _build_cell_formula(subfolder_path: str) -> str:
    r"""
    Baut eine portable CELL("filename")-Hyperlink-Formel.

    Die Formel berechnet den Hyperlink-Pfad dynamisch zur Laufzeit:
      LEFT(CELL("filename"), FIND("[", CELL("filename"))-1) -> Ordnerpfad der Excel-Datei
      & "Belege\\datei.pdf" -> relativer Unterordner + Dateiname

    Dadurch friert Excel den Pfad NICHT ein - er bleibt nach Ordner-Verschiebung gueltig.

    WICHTIG: Funktioniert nur wenn die Datei bereits gespeichert ist.
    In einer neuen, ungespeicherten Datei gibt CELL("filename") "" zurueck.

    subfolder_path: relativer Teilpfad nach dem Ordner der Excel-Datei
                    z.B. "Belege\\0012_Telekom.pdf"
    """
    # Backslashes in der Formel müssen als \\ (4 Backslashes) codiert werden,
    # damit Excel 2 Backslashes interpretiert → ein Backslash im Pfad
    escaped = subfolder_path.replace("\\", "\\\\")
    return (
        f'=HYPERLINK('
        f'LEFT(CELL("filename"),FIND("[",CELL("filename"))-1)&"{escaped}",'
        f'"{subfolder_path.split(chr(92))[-1]}")'
    )


def create_excel(documents, pdf_map, output_path, year_label,
                 unc_base=None, ocr_results=None, subfolder: str = "",
                 hyperlink_mode: str = "cell", include_text_path: bool = False,
                 progress_fn=None, overwrite: bool = False):
    """
    Erstellt die Excel-Datei im Steuerberater-Format (Stufe 1).

    documents:         Liste von Paperless-Dokumenten (API-Dicts)
    pdf_map:           {doc_id: filename_in_pdf_folder}
    output_path:       Zielpfad der .xlsx-Datei
    year_label:        z.B. "2024"
    unc_base:          Windows-UNC-Pfad Basis (optional)
    ocr_results:       {doc_id: {"absender": ..., "betrag": ...}} (optional, Stufe 2)
    subfolder:         optionaler Unterordner (Allowlist-validiert, default="")
    hyperlink_mode:    "cell" = CELL()-Formel (portabel); "unc" = absoluter UNC-Pfad (Issue #8)
    include_text_path: True = Spalte K mit kopierbarem UNC-Pfad (Issue #8)
    progress_fn:       optional callable(idx, total, title) – Fortschritt pro Zeile (Issue #19)
    overwrite:         False (Default): bestehende Datei nie überschreiben (manuelle Arbeit schützen)
    """
    if os.path.exists(output_path) and not overwrite:
        raise FileExistsError(
            f"Excel existiert bereits und wird nicht überschrieben "
            f"(manuelle Einträge schützen): {output_path}. "
            f"Bitte „Nur neue hinzufügen“ (Append) verwenden."
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rechnungsaufstellung"

    # ── Zeile 1: SUMME-Zeile ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    sum_label = ws.cell(row=1, column=1, value=f"SUMME {year_label}")
    sum_label.font      = Font(bold=True, size=11, color=COLOR_SUM_FONT)
    sum_label.fill      = PatternFill("solid", fgColor=COLOR_SUM_BG)
    sum_label.alignment = Alignment(horizontal="left", vertical="center")
    sum_label.border    = BORDER

    # ── Zeilen 2–3: leer ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 8

    # ── Zeile 4: Spaltenheader ───────────────────────────────────────────
    ws.row_dimensions[4].height = 36
    active_columns = list(COLUMNS)
    if include_text_path:
        active_columns.append(COLUMN_TEXT_PATH)
    for col_idx, (header, width, align) in enumerate(active_columns, start=1):
        _header_cell(ws, 4, col_idx, header, align)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Daten ab Zeile 5 ─────────────────────────────────────────────────
    data_start_row = 5
    sorted_docs = sorted(
        documents,
        key=lambda d: d.get("created", "1900-01-01") or "1900-01-01"
    )
    ocr = ocr_results or {}
    total_docs = len(sorted_docs)

    for i, doc in enumerate(sorted_docs):
        row = data_start_row + i
        ws.row_dimensions[row].height = 18
        doc_id = doc.get("id")
        if progress_fn:
            progress_fn(i + 1, total_docs, doc.get("title", f"Dokument {doc_id}"))

        # A: Beleg-Nr.
        beleg_nr = doc.get("archive_serial_number") or doc_id
        _data_cell(ws, row, 1, beleg_nr, align="center")

        # B: Re-Dat
        created_str = doc.get("created", "")
        dt = None
        if created_str:
            try:
                dt = datetime.strptime(created_str[:10], "%Y-%m-%d").date()
            except ValueError:
                pass
        _data_cell(ws, row, 2, dt, align="center", number_fmt=DATE_FORMAT)

        # C: Zahlungsdatum – leer, manuell
        _data_cell(ws, row, 3, None, align="center",
                   number_fmt=DATE_FORMAT, bg=COLOR_EMPTY_BG)

        # D: Bar / Konto / KK – leer, manuell
        _data_cell(ws, row, 4, None, align="center", bg=COLOR_EMPTY_BG)

        # E: Absender
        ocr_data    = ocr.get(doc_id, {})
        correspondent = doc.get("correspondent_name") or None
        ocr_absender  = ocr_data.get("absender")

        if correspondent:
            # Aus Paperless – zuverlässig
            _data_cell(ws, row, 5, correspondent, align="left")
        elif ocr_absender:
            # OCR-Vorschlag – gelb
            cell_e = _data_cell(ws, row, 5, ocr_absender, align="left",
                                 bg=COLOR_OCR_BG)
            cell_e.comment = _make_comment("OCR-Vorschlag – bitte prüfen!")
        else:
            _data_cell(ws, row, 5, None, align="left", bg=COLOR_EMPTY_BG)

        # F: Beschreibung
        _data_cell(ws, row, 6, doc.get("title", ""), align="left")

        # G: Kennzahl (Dokumenttyp)
        doc_type_name = (doc.get("document_type_name")
                         or str(doc.get("document_type", "")) or "")
        _data_cell(ws, row, 7, doc_type_name, align="left")

        # H: Rechnungssumme
        ocr_betrag = ocr_data.get("betrag")
        if ocr_betrag is not None:
            cell_h = _data_cell(ws, row, 8, ocr_betrag, align="right",
                                 number_fmt=NUMBER_FORMAT, bg=COLOR_OCR_BG)
            cell_h.comment = _make_comment("OCR-Vorschlag – bitte prüfen!")
        else:
            _data_cell(ws, row, 8, None, align="right",
                       number_fmt=NUMBER_FORMAT, bg=COLOR_EMPTY_BG)

        # I: Rechnungssumme inkl. Privatanteil – leer, manuell
        _data_cell(ws, row, 9, None, align="right",
                   number_fmt=NUMBER_FORMAT, bg=COLOR_EMPTY_BG)

        # J: Dateiname / Hyperlink (Issue #8: CELL-Formel oder UNC)
        filename = pdf_map.get(doc_id, "")
        if filename:
            if hyperlink_mode == "cell":
                # CELL("filename")-Formel: portabel, funktioniert nach Ordner-Verschiebung
                if subfolder:
                    rel_path = f"{subfolder}\\Belege\\{filename}"
                else:
                    rel_path = f"Belege\\{filename}"
                formula = _build_cell_formula(rel_path)
            elif unc_base:
                # Absoluter UNC-Pfad (Fallback / HYPERLINK_MODE=unc)
                unc_path = _build_unc_path(unc_base, year_label, filename, subfolder)
                formula  = f'=HYPERLINK("{unc_path}","{filename}")'
            else:
                formula = filename

            cell_j = ws.cell(row=row, column=10, value=formula)
            cell_j.font      = Font(size=10, color=COLOR_HYPERLINK, underline="single")
            cell_j.alignment = Alignment(horizontal="left", vertical="center")
            cell_j.border    = BORDER

            # Spalte K: kopierbarer UNC-Pfad (INCLUDE_TEXT_PATH=true)
            if include_text_path and unc_base:
                unc_path = _build_unc_path(unc_base, year_label, filename, subfolder)
                cell_k = ws.cell(row=row, column=11, value=unc_path)
                cell_k.font      = Font(size=9, color="666666")
                cell_k.alignment = Alignment(horizontal="left", vertical="center")
                cell_k.border    = BORDER
        else:
            _data_cell(ws, row, 10, filename, align="left")

    last_data_row = data_start_row + len(sorted_docs) - 1

    # ── Excel-Tabelle (Tabelle1) ──────────────────────────────────────────
    num_cols  = len(COLUMNS) + (1 if include_text_path else 0)
    table_ref = f"A4:{get_column_letter(num_cols)}{last_data_row}"
    table = Table(displayName="Tabelle1", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,  showColumnStripes=False,
    )
    ws.add_table(table)

    # Summenformel (Spalte H = Rechnungssumme)
    ws["H1"] = f"=SUM(H{data_start_row}:H{last_data_row})"
    ws["H1"].number_format = NUMBER_FORMAT
    ws["H1"].font          = Font(bold=True, size=11, color=COLOR_SUM_FONT)
    ws["H1"].fill          = PatternFill("solid", fgColor=COLOR_SUM_BG)
    ws["H1"].alignment     = Alignment(horizontal="right", vertical="center")
    ws["H1"].border        = BORDER

    ws.freeze_panes = "A5"
    ws.print_area   = f"A1:{get_column_letter(len(COLUMNS))}{last_data_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True

    wb.save(output_path)
    return output_path


def _cell_fill_is(cell, hex6: str) -> bool:
    """Vergleicht Zellenfarbe robust (mit/ohne Alpha-Prefix)."""
    try:
        fill = cell.fill
        if not fill or not fill.fgColor:
            return False
        rgb = fill.fgColor.rgb
        if rgb is None:
            return False
        return str(rgb).upper().endswith(str(hex6).upper())
    except Exception:
        return False


def update_excel_with_ocr(excel_path, ocr_results, unc_base, year_label,
                          subfolder: str = ""):
    """
    Stufe 2: Öffnet bestehendes Excel und trägt OCR-Ergebnisse ein.
    Überschreibt nur leere oder bereits gelbe Felder (schützt manuelle Einträge).

    ocr_results: {doc_id: {"absender": str|None, "betrag": float|None}}
    subfolder:   optionaler Unterordner unterhalb des Jahres-Ordners (default="")
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel nicht gefunden: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = _get_invoice_sheet(wb)

    YELLOW_FILL = PatternFill("solid", fgColor=COLOR_OCR_BG)
    EMPTY_FILL  = PatternFill("solid", fgColor=COLOR_EMPTY_BG)
    NONE_FILL   = PatternFill(fill_type=None)

    updated = 0

    # Daten starten ab Zeile 5 – Spalte A = Beleg-Nr. (nutzen wir als Key)
    # Wir brauchen eine Zuordnung Beleg-Nr./ID → Zeile
    # Da doc_id in Spalte A steht (ASN oder ID), iterieren wir

    # Baue Mapping: Wert in Spalte A → Zeile
    row_map = {}
    for row in ws.iter_rows(min_row=5):
        cell_a = row[0]
        if cell_a.value is not None:
            row_map[str(cell_a.value)] = cell_a.row

    for doc_id, data in ocr_results.items():
        # doc_id könnte ASN oder numerische ID sein – probiere beide
        row_num = row_map.get(str(doc_id))
        if row_num is None:
            continue

        absender = data.get("absender")
        betrag   = data.get("betrag")

        # Spalte E (Absender) – nur überschreiben wenn leer oder gelb
        cell_e = ws.cell(row=row_num, column=5)
        e_is_empty = cell_e.value is None or cell_e.value == ""
        e_is_ocr   = _cell_fill_is(cell_e, COLOR_OCR_BG)
        if absender and (e_is_empty or e_is_ocr):
            cell_e.value      = absender
            cell_e.fill       = YELLOW_FILL
            cell_e.font       = Font(size=10)
            cell_e.alignment  = Alignment(horizontal="left", vertical="center")
            cell_e.border     = BORDER
            cell_e.comment    = _make_comment("OCR-Vorschlag – bitte prüfen!")
            updated += 1

        # Spalte H (Rechnungssumme) – nur überschreiben wenn leer oder gelb
        cell_h = ws.cell(row=row_num, column=8)
        h_is_empty = cell_h.value is None or cell_h.value == ""
        h_is_ocr   = _cell_fill_is(cell_h, COLOR_OCR_BG)
        if betrag is not None and (h_is_empty or h_is_ocr):
            cell_h.value         = betrag
            cell_h.fill          = YELLOW_FILL
            cell_h.font          = Font(size=10)
            cell_h.alignment     = Alignment(horizontal="right", vertical="center")
            cell_h.number_format = NUMBER_FORMAT
            cell_h.border        = BORDER
            cell_h.comment       = _make_comment("OCR-Vorschlag – bitte prüfen!")
            updated += 1

        # Spalte J (Hyperlink) – nur Plain-Filename → HYPERLINK; CELL()-Formeln unangetastet
        cell_j = ws.cell(row=row_num, column=10)
        raw_j = cell_j.value
        if raw_j is None or raw_j == "":
            continue
        filename = str(raw_j)
        if filename.startswith("="):
            continue  # bestehende Formel (CELL/HYPERLINK) nicht umschreiben
        if unc_base:
            unc_path = _build_unc_path(unc_base, year_label, filename, subfolder)
            cell_j.value      = f'=HYPERLINK("{unc_path}","{filename}")'
            cell_j.font       = Font(size=10, color=COLOR_HYPERLINK, underline="single")
            cell_j.alignment  = Alignment(horizontal="left", vertical="center")
            cell_j.border     = BORDER

    wb.save(excel_path)
    return updated


def get_existing_doc_ids(excel_path):
    """
    Liest alle Beleg-Nummern (Spalte A ab Zeile 5) aus einem bestehenden Excel.
    Gibt ein Set von Strings zurück (ASN oder numerische ID).
    """
    if not os.path.exists(excel_path):
        return set()
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = _get_invoice_sheet(wb)
    ids = set()
    for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            ids.add(str(val))
    wb.close()
    return ids


def append_to_excel(new_documents, pdf_map, excel_path, year_label,
                    unc_base=None, subfolder: str = "",
                    hyperlink_mode: str = "cell", include_text_path: bool = False):
    """
    Hängt neue Dokumente an ein bestehendes Excel an.
    Bestehende Zeilen werden nicht verändert.
    Gibt die Anzahl neu angehängter Zeilen zurück.

    new_documents:     Liste von Paperless-Dokumenten die noch NICHT im Excel sind
    pdf_map:           {doc_id: filename}
    subfolder:         optionaler Unterordner (default="")
    hyperlink_mode:    "cell" = CELL()-Formel; "unc" = absoluter UNC-Pfad (Issue #8)
    include_text_path: Spalte K mit kopierbarem UNC-Pfad (Issue #8)
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel nicht gefunden: {excel_path}")
    if not new_documents:
        return 0

    wb = openpyxl.load_workbook(excel_path)
    ws = _get_invoice_sheet(wb)

    # Letzte belegte Zeile in Spalte A finden
    last_row = 4  # Mindestens Header-Zeile
    for row in ws.iter_rows(min_row=5, max_col=1):
        if row[0].value is not None:
            last_row = row[0].row

    insert_start = last_row + 1

    sorted_new = sorted(
        new_documents,
        key=lambda d: d.get("created", "1900-01-01") or "1900-01-01"
    )

    for i, doc in enumerate(sorted_new):
        row = insert_start + i
        ws.row_dimensions[row].height = 18
        doc_id = doc.get("id")

        # A: Beleg-Nr.
        beleg_nr = doc.get("archive_serial_number") or doc_id
        _data_cell(ws, row, 1, beleg_nr, align="center")

        # B: Re-Dat
        created_str = doc.get("created", "")
        dt = None
        if created_str:
            try:
                from datetime import datetime as _dt
                dt = _dt.strptime(created_str[:10], "%Y-%m-%d").date()
            except ValueError:
                pass
        _data_cell(ws, row, 2, dt, align="center", number_fmt=DATE_FORMAT)

        # C: Zahlungsdatum
        _data_cell(ws, row, 3, None, align="center",
                   number_fmt=DATE_FORMAT, bg=COLOR_EMPTY_BG)
        # D: Bar / Konto / KK
        _data_cell(ws, row, 4, None, align="center", bg=COLOR_EMPTY_BG)

        # E: Absender
        correspondent = doc.get("correspondent_name") or None
        if correspondent:
            _data_cell(ws, row, 5, correspondent, align="left")
        else:
            _data_cell(ws, row, 5, None, align="left", bg=COLOR_EMPTY_BG)

        # F: Beschreibung
        _data_cell(ws, row, 6, doc.get("title", ""), align="left")

        # G: Kennzahl
        doc_type_name = (doc.get("document_type_name")
                         or str(doc.get("document_type", "")) or "")
        _data_cell(ws, row, 7, doc_type_name, align="left")

        # H: Rechnungssumme (leer – kann später per Stufe 2 befüllt werden)
        _data_cell(ws, row, 8, None, align="right",
                   number_fmt=NUMBER_FORMAT, bg=COLOR_EMPTY_BG)

        # I: Rechnungssumme inkl. Privatanteil
        _data_cell(ws, row, 9, None, align="right",
                   number_fmt=NUMBER_FORMAT, bg=COLOR_EMPTY_BG)

        # J: Dateiname / Hyperlink (Issue #8)
        filename = pdf_map.get(doc_id, "")
        if filename:
            if hyperlink_mode == "cell":
                if subfolder:
                    rel_path = f"{subfolder}\\Belege\\{filename}"
                else:
                    rel_path = f"Belege\\{filename}"
                formula = _build_cell_formula(rel_path)
            elif unc_base:
                unc_path = _build_unc_path(unc_base, year_label, filename, subfolder)
                formula  = f'=HYPERLINK("{unc_path}","{filename}")'
            else:
                formula = filename

            cell_j = ws.cell(row=row, column=10, value=formula)
            cell_j.font      = Font(size=10, color=COLOR_HYPERLINK, underline="single")
            cell_j.alignment = Alignment(horizontal="left", vertical="center")
            cell_j.border    = BORDER

            if include_text_path and unc_base:
                unc_path = _build_unc_path(unc_base, year_label, filename, subfolder)
                cell_k = ws.cell(row=row, column=11, value=unc_path)
                cell_k.font      = Font(size=9, color="666666")
                cell_k.alignment = Alignment(horizontal="left", vertical="center")
                cell_k.border    = BORDER
        else:
            _data_cell(ws, row, 10, filename, align="left")

    # SUMME-Formel auf neue letzte Zeile ausdehnen – nur eigenes SUM, keine manuellen H1-Formeln
    new_last_row = insert_start + len(sorted_new) - 1
    h1 = ws["H1"]
    h1_val = h1.value
    if h1_val is None or (
        isinstance(h1_val, str) and h1_val.strip().upper().startswith("=SUM")
    ):
        ws["H1"] = f"=SUM(H5:H{new_last_row})"
        ws["H1"].number_format = NUMBER_FORMAT
        ws["H1"].font          = Font(bold=True, size=11, color=COLOR_SUM_FONT)
        ws["H1"].fill          = PatternFill("solid", fgColor=COLOR_SUM_BG)
        ws["H1"].alignment     = Alignment(horizontal="right", vertical="center")
        ws["H1"].border        = BORDER

    # Tabellen-Zeilen erweitern (Spaltenbreite behalten, nie schrumpfen) + Columns syncen
    _expand_invoice_table_rows(ws, new_last_row, include_text_path=include_text_path)

    wb.save(excel_path)
    return len(sorted_new)


def _expand_invoice_table_rows(ws, new_last_row: int, *, include_text_path: bool = False) -> None:
    """
    Erweitert die Rechnungs-Tabelle nach unten. Spaltenbreite bleibt mindestens
    der bisherige Stand (kein Schrumpfen auf len(COLUMNS) – schont Match-Spalten).
    """
    from openpyxl.utils import range_boundaries

    tbl = _find_invoice_table(ws)
    if tbl is None:
        return
    try:
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        needed = 11 if include_text_path else len(COLUMNS)
        last_col = max(max_col, needed)
        if new_last_row < max_row and last_col == max_col:
            return
        end_row = max(max_row, new_last_row)
        tbl.ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(last_col)}{end_row}"
        )
        _sync_table_columns_from_header(ws, tbl)
    except Exception:
        return


# ── Issue #25: Kontoauszug-Matching ───────────────────────────────────

def _header_map(ws) -> dict:
    """Header-Zeile 4 → {normalisierter Name: Spaltenindex 1-basiert}."""
    result = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val is None:
            continue
        key = str(val).replace("\n", " ").strip().lower()
        result[key] = col
    return result


def _ensure_bank_match_columns(ws) -> tuple:
    """
    Stellt sicher, dass Spalten Buchungstext und Match-Status existieren.
    Neue Spalten werden direkt rechts an die Rechnungs-Tabelle gehängt
    (nicht hinter fremde Spalten/Tabellen). Issue #32 / Excel-Repair.
    Rückgabe: (col_booking_text, col_match_status)
    """
    from openpyxl.utils import range_boundaries

    headers = _header_map(ws)
    col_text = None
    col_status = None
    for name, col in headers.items():
        if "buchungstext" in name:
            col_text = col
        if "match-status" in name or name == "match status":
            col_status = col

    tbl = _find_invoice_table(ws)
    if tbl is not None:
        try:
            _min_c, _min_r, max_col, _max_r = range_boundaries(tbl.ref)
            next_col = max_col + 1
        except Exception:
            next_col = None
    else:
        next_col = None

    if next_col is None:
        last_header = 0
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=4, column=col).value is not None:
                last_header = col
        next_col = last_header + 1

    added = False

    def _claim_header_col(preferred: int) -> int:
        """Nächste freie Header-Zelle ab preferred (keine Übernahme fremder Inhalte)."""
        col = preferred
        for _ in range(50):
            val = ws.cell(row=4, column=col).value
            if val is None or str(val).strip() == "":
                return col
            col += 1
        return preferred

    if col_text is None:
        col_text = _claim_header_col(next_col)
        _header_cell(ws, 4, col_text, COLUMN_BOOKING_TEXT[0], COLUMN_BOOKING_TEXT[2])
        ws.column_dimensions[get_column_letter(col_text)].width = COLUMN_BOOKING_TEXT[1]
        next_col = col_text + 1
        added = True
    if col_status is None:
        col_status = _claim_header_col(next_col)
        _header_cell(ws, 4, col_status, COLUMN_MATCH_STATUS[0], COLUMN_MATCH_STATUS[2])
        ws.column_dimensions[get_column_letter(col_status)].width = COLUMN_MATCH_STATUS[1]
        added = True

    # Tabelle nur erweitern, wenn wirklich neue Spalten angelegt wurden.
    # Reines Sync bei jedem Schreiben zerstört strukturierte Formeln (#BEZUG!).
    if added:
        _expand_table_columns_only(ws, max(col_text, col_status))

    return col_text, col_status


def _is_formula_cell(cell) -> bool:
    v = cell.value
    return isinstance(v, str) and v.startswith("=")


def _cell_is_empty(cell) -> bool:
    """True wenn kein nutzbarer Wert (None, '', Whitespace)."""
    v = cell.value
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _cell_is_writable(cell) -> bool:
    """Leer und keine Formel → Wert darf geschrieben werden."""
    if _is_formula_cell(cell):
        return False
    return _cell_is_empty(cell)


# Bekannte App-Statuswerte – dürfen bei Re-Run aktualisiert werden
_MATCH_STATUS_VALUES = frozenset({
    "offen", "gefunden", "mehrdeutig", "nicht gefunden", "kein Betrag",
})


def _status_cell_writable(cell) -> bool:
    """Match-Status: leer oder bereits App-Status (kein manueller Freitext)."""
    if _is_formula_cell(cell):
        return False
    if _cell_is_empty(cell):
        return True
    return str(cell.value).strip() in _MATCH_STATUS_VALUES


def _find_invoice_table(ws):
    """
    Liefert die Rechnungs-Tabelle (oder None).
    Preferenz: displayName Tabelle1, sonst Tabelle die Header-Zeile 4 abdeckt.
    """
    from openpyxl.utils import range_boundaries

    tables = list(ws.tables.values())
    if not tables:
        return None
    for tbl in tables:
        if (tbl.displayName or "").strip().lower() in ("tabelle1", "table1"):
            return tbl
    for tbl in tables:
        try:
            _c1, r1, _c2, r2 = range_boundaries(tbl.ref)
            if r1 <= 4 <= r2:
                return tbl
        except Exception:
            continue
    return tables[0]


def _sync_table_columns_from_header(ws, tbl) -> None:
    """
    tableColumns + autoFilter an tbl.ref anpassen.

    Wichtig für strukturierte Formeln (sonst #BEZUG! / Excel-Repair):
      - bestehende tableColumn-Namen unverändert lassen
      - Header-Zellen nie umbenennen / kein \\n→Leerzeichen
      - nur für neu hinzugekommene Spalten Namen aus dem Header übernehmen
        (exakter Zelltext inkl. Zeilenumbruch)
    """
    from openpyxl.utils import range_boundaries
    from openpyxl.worksheet.table import TableColumn

    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
    needed = max_col - min_col + 1
    existing = list(tbl.tableColumns) if tbl.tableColumns else []

    # Schon konsistent → nichts anfassen (verhindert Excel-Repair)
    if len(existing) == needed:
        if tbl.autoFilter is not None:
            tbl.autoFilter.ref = tbl.ref
        return

    used_names = set()
    columns = []

    # Bestehende Spalten 1:1 behalten (Namen unverändert)
    for idx, old in enumerate(existing, start=1):
        if idx > needed:
            break
        name = old.name
        used_names.add(str(name).lower())
        columns.append(TableColumn(id=idx, name=name))

    # Nur neue Spalten rechts anhängen
    for idx, col in enumerate(range(min_col, max_col + 1), start=1):
        if idx <= len(columns):
            continue
        raw = ws.cell(row=min_row, column=col).value
        # Exakter Header-Text – Zeilenumbrüche erhalten (strukturierte Bezüge)
        if raw not in (None, ""):
            name = str(raw)
        else:
            name = f"Spalte{idx}"
            # Nur leere Header füllen – nie bestehende umbenennen
            ws.cell(row=min_row, column=col).value = name
        base = name
        n = 2
        while name.lower() in used_names:
            name = f"{base} ({n})"
            n += 1
        used_names.add(name.lower())
        columns.append(TableColumn(id=idx, name=name))

    tbl.tableColumns = columns
    if tbl.autoFilter is not None:
        tbl.autoFilter.ref = tbl.ref
    else:
        try:
            from openpyxl.worksheet.filters import AutoFilter
            tbl.autoFilter = AutoFilter(ref=tbl.ref)
        except Exception:
            pass


def _table_ranges_overlap(ref_a: str, ref_b: str) -> bool:
    from openpyxl.utils import range_boundaries
    a1, r1, a2, r2 = range_boundaries(ref_a)
    b1, s1, b2, s2 = range_boundaries(ref_b)
    return not (a2 < b1 or b2 < a1 or r2 < s1 or s2 < r1)


def _expand_table_columns_only(ws, last_col: int) -> None:
    """
    Erweitert NUR die Rechnungs-Tabelle nach rechts (Spalten).
    Zeilenbereich unverändert. Andere Tabellen unangetastet.
    Wenn last_col bereits in tbl.ref liegt: kein Sync, keine Mutation.
    """
    from openpyxl.utils import range_boundaries

    tbl = _find_invoice_table(ws)
    if tbl is None:
        return

    try:
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
    except Exception:
        return

    if last_col <= max_col:
        # Spalten schon im Ref – Tabelle nicht anfassen (schützt Formeln / #BEZUG!)
        return

    new_ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(last_col)}{max_row}"
    )

    for other in ws.tables.values():
        if other is tbl or other.name == tbl.name:
            continue
        try:
            if _table_ranges_overlap(new_ref, other.ref):
                return
        except Exception:
            continue

    try:
        tbl.ref = new_ref
        _sync_table_columns_from_header(ws, tbl)
    except Exception:
        return


def _parse_excel_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _cell_as_float(value):
    """Zahl aus Zelle; Formeln/leer → None (openpyxl data_only ohne Cache)."""
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.strip().startswith("="):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _cell_is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _beleg_nr_missing(beleg) -> bool:
    """True wenn Spalte A keine nutzbare Beleg-Nr. hat."""
    if _cell_is_blank(beleg):
        return True
    if isinstance(beleg, str) and beleg.strip() in ("-", "–", "—", "?"):
        return True
    return False


def _beleg_as_int(beleg) -> Optional[int]:
    """Numerische Beleg-Nr. (84, 100, '0001'); Formeln/leer → None."""
    if isinstance(beleg, str) and beleg.strip().startswith("="):
        return None
    if _beleg_nr_missing(beleg):
        return None
    if isinstance(beleg, bool):
        return None
    if isinstance(beleg, (int, float)):
        return int(beleg)
    s = str(beleg).strip().replace(" ", "")
    if re.fullmatch(r"\d+", s):
        return int(s)
    return None


BELEG_AUTO_COMMENT = "Automatisch vergeben – bitte prüfen!"


def _row_has_invoice_content(ws, row: int) -> bool:
    """True wenn die Zeile wie eine Rechnung aussieht (nicht nur Padding)."""
    beleg = ws.cell(row=row, column=1).value
    re_raw = ws.cell(row=row, column=2).value
    absender = ws.cell(row=row, column=5).value
    beschreibung = ws.cell(row=row, column=6).value
    absender_s = str(absender).strip() if absender else ""
    beschreibung_s = str(beschreibung).strip() if beschreibung else ""
    h = _cell_as_float(ws.cell(row=row, column=8).value)
    i = _cell_as_float(ws.cell(row=row, column=9).value)
    return (
        (not _cell_is_blank(beleg) and not (isinstance(beleg, str) and beleg.startswith("=")))
        or _parse_excel_date(re_raw) is not None
        or bool(absender_s)
        or bool(beschreibung_s)
        or h is not None
        or i is not None
    )


def _max_existing_beleg_nr(ws) -> int:
    mx = 0
    for row in range(5, ws.max_row + 1):
        n = _beleg_as_int(ws.cell(row=row, column=1).value)
        if n is not None and n > mx:
            mx = n
    return mx


def _beleg_cell_assignable(cell) -> bool:
    """
    Wie Matching-Schreibschutz: nur wirklich leere Zellen, keine Formeln,
    keine vorhandenen Werte (auch keine Platzhalter).
    """
    return _cell_is_writable(cell)


def plan_auto_beleg_numbers(excel_path) -> list:
    """
    Plant fortlaufende Beleg-Nrn. für Inhaltszeilen mit leerer Spalte A.
    Start = max(vorhandene numerische Beleg-Nrn.) + 1.
    Schreibt nicht. Formeln und befüllte Zellen werden übersprungen.
    Rückgabe: [{row, proposed_beleg_nr, absender, ...}, ...]
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel nicht gefunden: {excel_path}")

    # Ohne data_only: Formeln in A erkennen (nicht überschreiben)
    wb = openpyxl.load_workbook(excel_path)
    ws = _get_invoice_sheet(wb)
    next_nr = _max_existing_beleg_nr(ws) + 1
    max_existing = next_nr - 1
    planned = []
    for row in range(5, ws.max_row + 1):
        if not _row_has_invoice_content(ws, row):
            continue
        cell_a = ws.cell(row=row, column=1)
        if not _beleg_cell_assignable(cell_a):
            continue
        absender = ws.cell(row=row, column=5).value
        beschreibung = ws.cell(row=row, column=6).value
        planned.append({
            "row": row,
            "beleg_nr": None,
            "proposed_beleg_nr": next_nr,
            "re_dat": _parse_excel_date(ws.cell(row=row, column=2).value),
            "absender": str(absender).strip() if absender else "",
            "beschreibung": str(beschreibung).strip() if beschreibung else "",
            "betrag": (
                _cell_as_float(ws.cell(row=row, column=9).value)
                or _cell_as_float(ws.cell(row=row, column=8).value)
            ),
            "code": "beleg_auto_assign",
            "message": (
                f"Zeile {row}: Beleg-Nr. → {next_nr} "
                f"(automatisch, bitte prüfen)"
            ),
            "max_existing_beleg_nr": max_existing,
        })
        next_nr += 1
    wb.close()
    return planned


def apply_auto_beleg_numbers(excel_path, planned: list | None = None) -> list:
    """
    Vergibt Beleg-Nrn. in Spalte A (gelb + „bitte prüfen“).

    Schreibschutz (wie Bank-Match):
      - nur leere Zellen (keine vorhandenen Werte/Platzhalter)
      - keine Formeln
      - Nummern werden live aus max(A)+1 neu berechnet (kein Blind-Schreiben
        aus veralteter Vorschau)
      - bestehende Füllungen/Kommentare auf nicht-leeren Zellen bleiben unberührt
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel nicht gefunden: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = _get_invoice_sheet(wb)

    # Zeilenreihenfolge: Vorschau beibehalten, sonst Sheet-Scan
    if planned:
        candidate_rows = [p["row"] for p in planned if p.get("row") is not None]
    else:
        candidate_rows = [
            row for row in range(5, ws.max_row + 1)
            if _row_has_invoice_content(ws, row)
        ]

    next_nr = _max_existing_beleg_nr(ws) + 1
    applied = []
    for row in candidate_rows:
        cell = ws.cell(row=row, column=1)
        if not _beleg_cell_assignable(cell):
            continue
        if not _row_has_invoice_content(ws, row):
            continue

        cell.value = next_nr
        cell.fill = PatternFill("solid", fgColor=COLOR_OCR_BG)
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        cell.comment = _make_comment(BELEG_AUTO_COMMENT)

        absender = ws.cell(row=row, column=5).value
        beschreibung = ws.cell(row=row, column=6).value
        applied.append({
            "row": row,
            "beleg_nr": next_nr,
            "proposed_beleg_nr": next_nr,
            "re_dat": _parse_excel_date(ws.cell(row=row, column=2).value),
            "absender": str(absender).strip() if absender else "",
            "beschreibung": str(beschreibung).strip() if beschreibung else "",
            "betrag": (
                _cell_as_float(ws.cell(row=row, column=9).value)
                or _cell_as_float(ws.cell(row=row, column=8).value)
            ),
            "code": "beleg_auto_assign",
            "message": (
                f"Zeile {row}: Beleg-Nr. → {next_nr} "
                f"(automatisch, bitte prüfen)"
            ),
        })
        next_nr += 1

    if applied:
        wb.save(excel_path)
    wb.close()
    return applied


def collect_missing_beleg_issues(invoices: list, planned: list | None = None) -> list:
    """
    Datenqualität / Auto-Vergabe-Vorschau für Zeilen ohne Beleg-Nr.
    planned: optional plan_auto_beleg_numbers()-Ergebnis (proposed_beleg_nr).
    """
    by_row = {p["row"]: p for p in (planned or [])}
    issues = []
    for inv in invoices:
        if not _beleg_nr_missing(inv.get("beleg_nr")):
            continue
        prop = by_row.get(inv.get("row"), {})
        proposed = prop.get("proposed_beleg_nr")
        issues.append({
            "row": inv.get("row"),
            "beleg_nr": inv.get("beleg_nr"),
            "proposed_beleg_nr": proposed,
            "re_dat": inv.get("re_dat"),
            "absender": inv.get("absender") or "",
            "beschreibung": inv.get("beschreibung") or "",
            "betrag": inv.get("betrag"),
            "code": "beleg_auto_assign",
            "message": (
                f"Zeile {inv.get('row')}: Beleg-Nr. → {proposed} "
                f"(automatisch, bitte prüfen)"
                if proposed is not None
                else (
                    f"Zeile {inv.get('row')}: Beleg-Nr. (Spalte A) fehlt – "
                    f"{(inv.get('absender') or inv.get('beschreibung') or 'ohne Absender')}"
                )
            ),
        })
    return issues


def read_invoices_for_matching(excel_path) -> list:
    """
    Liest Rechnungszeilen mit **leerem Zahlungsdatum (C)** fürs Matching.
    Rückgabe: [{row, beleg_nr, re_dat, absender, beschreibung, betrag}, ...]

    Zeilen ohne Beleg-Nr. (A) werden mitgenommen, wenn sonst Inhalt da ist
    (Re-Dat / Absender / Beschreibung / Betrag) – in bearbeiteten Dateien
    fehlen Belegnummern oft, obwohl die Rechnung vollständig ist.
    Reine Padding-Zeilen (alles leer) werden übersprungen.
    Beim Schreiben vergibt apply_auto_beleg_numbers() fortlaufende Nummern
    (gelb + „bitte prüfen“).

    Betrag für Bank-Abgleich:
      Spalte I (Rechnungssumme inkl. Privatanteil), falls gesetzt –
      die Bank bucht den Vollbetrag. Sonst Spalte H (Rechnungssumme).
      H ist oft eine Anteil-Formel (z. B. I*0.11); ohne Excel-Cache liefert
      data_only dafür None – I bleibt dann die einzige nutzbare Zahl.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel nicht gefunden: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = _get_invoice_sheet(wb)
    invoices = []
    for row in range(5, ws.max_row + 1):
        zahlung = ws.cell(row=row, column=3).value
        if not _cell_is_blank(zahlung):
            continue  # Entscheidung #2: nur leeres C

        beleg = ws.cell(row=row, column=1).value
        re_dat = _parse_excel_date(ws.cell(row=row, column=2).value)
        absender = ws.cell(row=row, column=5).value
        beschreibung = ws.cell(row=row, column=6).value
        absender_s = str(absender).strip() if absender else ""
        beschreibung_s = str(beschreibung).strip() if beschreibung else ""
        # Bank = Vollbetrag (I), steuerlicher Anteil (H) nur als Fallback
        betrag_f = _cell_as_float(ws.cell(row=row, column=9).value)
        if betrag_f is None:
            betrag_f = _cell_as_float(ws.cell(row=row, column=8).value)

        has_content = (
            not _cell_is_blank(beleg)
            or re_dat is not None
            or bool(absender_s)
            or bool(beschreibung_s)
            or betrag_f is not None
        )
        if not has_content:
            continue  # Padding / leere Tabellenzeile

        invoices.append({
            "row": row,
            "beleg_nr": beleg,
            "re_dat": re_dat,
            "absender": absender_s,
            "beschreibung": beschreibung_s,
            "betrag": betrag_f,
        })
    wb.close()
    return invoices


def _clear_cell_fill_comment(cell) -> None:
    cell.fill = PatternFill(fill_type=None)
    cell.comment = None


def _style_payment_cell(cell, fill_color, comment=None):
    """Farbe/Kommentar auf Spalte C (führender Indikator, Issue #34)."""
    if _is_formula_cell(cell):
        return False
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    if comment:
        cell.comment = _make_comment(comment)
    return True


def update_excel_with_bank_matches(excel_path, match_result: dict) -> int:
    """
    Schreibt Matching-Ergebnis NUR in die Rechnungs-Excel (Issue #32 / #34).

    Getrennte Zellen:
      - Spalte C: nur Zahlungsdatum (+ Farbe/Kommentar)
      - Spalte Buchungstext: nur Buchungstext
      - Spalte Match-Status: nur Status

    Schreibschutz: Werte nur in zuvor leere Zellen (keine Formeln).
    Formeln und Tabellen-Definition bleiben erhalten (kein unnötiges tableColumns-Rewrite).
    """
    from matching_csv import (
        STATUS_FOUND, STATUS_AMBIGUOUS, STATUS_NOT_FOUND, STATUS_NO_AMOUNT,
    )

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel nicht gefunden: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = _get_invoice_sheet(wb)
    col_text, col_status = _ensure_bank_match_columns(ws)

    # Harte Trennung: Datum nie in Buchungstext-/Status-Spalte und umgekehrt
    if col_text in (3, col_status) or col_status == 3:
        wb.close()
        raise ValueError(
            f"Ungültige Match-Spalten (C={3}, Buchungstext={col_text}, "
            f"Match-Status={col_status}) – Abbruch ohne Schreiben."
        )

    updated = 0

    def _set_status(row, status, fill_color, comment=None):
        nonlocal updated
        cell = ws.cell(row=row, column=col_status)
        if not _status_cell_writable(cell):
            return
        cell.value = status
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        if comment:
            cell.comment = _make_comment(comment)
        updated += 1

    def _set_text(row, text, fill_color):
        nonlocal updated
        cell = ws.cell(row=row, column=col_text)
        if not _cell_is_writable(cell):
            return
        # Nur Buchungstext – kein Datum in dieser Zelle
        cell.value = text or ""
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
        updated += 1

    def _mark_empty_c(row, fill_color, comment=None):
        """Farbe/Kommentar auf C nur wenn leer (kein Wert / kein Buchungstext)."""
        nonlocal updated
        cell_c = ws.cell(row=row, column=3)
        if not _cell_is_writable(cell_c):
            return
        _style_payment_cell(cell_c, fill_color, comment)
        updated += 1

    for m in match_result.get("matches", []):
        row = m["invoice_row"]
        cell_c = ws.cell(row=row, column=3)
        if not _cell_is_writable(cell_c):
            # C belegt/Formel → weder Datum noch Hilfsspalten anfassen
            continue
        # Spalte C: ausschließlich Datum
        cell_c.value = m["date"]
        cell_c.number_format = DATE_FORMAT
        _style_payment_cell(
            cell_c, COLOR_MATCH_OK,
            "Aus Kontoauszug zugeordnet – bitte prüfen!",
        )
        updated += 1
        # Spalte Buchungstext: ausschließlich Text (andere Zelle)
        _set_text(row, m.get("text") or "", COLOR_MATCH_OK)
        _set_status(row, STATUS_FOUND, COLOR_MATCH_OK)

    for a in match_result.get("ambiguous", []):
        row = a["row"]
        cands = a.get("candidates") or []
        lines = []
        for c in cands[:3]:
            lines.append(
                f"#{c.get('bank_row_id')} {c.get('date')} {c.get('amount')} "
                f"{(c.get('partner') or '')[:40]}"
            )
        comment = "Mehrdeutig:\n" + "\n".join(lines) if lines else "Mehrere Treffer"
        _mark_empty_c(row, COLOR_MATCH_AMBIG, comment)
        _set_status(row, STATUS_AMBIGUOUS, COLOR_MATCH_AMBIG, comment)

    for u in match_result.get("unmatched", []):
        row = u["row"]
        _mark_empty_c(row, COLOR_MATCH_MISS, "Nicht im Kontoauszug gefunden")
        _set_status(row, STATUS_NOT_FOUND, COLOR_MATCH_MISS)

    for n in match_result.get("no_amount", []):
        row = n["row"]
        _mark_empty_c(
            row, COLOR_MATCH_NO_AMT,
            "Kein Rechnungsbetrag (Spalte H) – Matching nicht möglich",
        )
        _set_status(row, STATUS_NO_AMOUNT, COLOR_MATCH_NO_AMT)

    wb.save(excel_path)
    return updated


def prepare_stb_export(excel_path) -> dict:
    """
    STB-Export-Modus (Issue #34): behält gefüllte Zahlungsdaten,
    entfernt Match-Farben/Kommentare und leert die Hilfsspalten Match-Status
    (und Buchungstext-Füllfarbe). Spalte Match-Status wird ausgeblendet.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel nicht gefunden: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = _get_invoice_sheet(wb)

    headers = {
        (ws.cell(row=4, column=c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=4, column=c).value
    }
    col_text = headers.get(COLUMN_BOOKING_TEXT[0])
    col_status = headers.get(COLUMN_MATCH_STATUS[0])

    cleared_c = 0
    cleared_status = 0
    cleared_text = 0
    cleared_beleg_auto = 0

    for row in range(5, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        if cell_a.value in (None, ""):
            continue
        # Auto-Beleg-Markierung entfernen (Nummer bleibt)
        cmt = cell_a.comment
        if cmt and BELEG_AUTO_COMMENT in (cmt.text or ""):
            _clear_cell_fill_comment(cell_a)
            cleared_beleg_auto += 1

        cell_c = ws.cell(row=row, column=3)
        if not _is_formula_cell(cell_c):
            had_fill = cell_c.fill and cell_c.fill.fill_type is not None
            had_comment = cell_c.comment is not None
            _clear_cell_fill_comment(cell_c)
            if had_fill or had_comment:
                cleared_c += 1

        if col_status:
            cell_s = ws.cell(row=row, column=col_status)
            if not _is_formula_cell(cell_s):
                if cell_s.value not in (None, "") or (
                    cell_s.fill and cell_s.fill.fill_type
                ) or cell_s.comment:
                    cell_s.value = None
                    _clear_cell_fill_comment(cell_s)
                    cleared_status += 1

        if col_text:
            cell_t = ws.cell(row=row, column=col_text)
            if not _is_formula_cell(cell_t):
                if cell_t.fill and cell_t.fill.fill_type:
                    _clear_cell_fill_comment(cell_t)
                    cleared_text += 1

    if col_status:
        ws.column_dimensions[get_column_letter(col_status)].hidden = True

    wb.save(excel_path)
    return {
        "cleared_payment_styles": cleared_c,
        "cleared_status": cleared_status,
        "cleared_text_styles": cleared_text,
        "cleared_beleg_auto_styles": cleared_beleg_auto,
        "status_column_hidden": bool(col_status),
    }
